# -*- coding: utf-8 -*-
"""一次性：从 parsed.json + bc_parsed.json 生成《master_bank.xlsx》初版。
母题库是清洗后的主工作副本（唯一人工编辑入口）；Word 原文为终极真值，
校对时以 Word 为准修正母题库。某部门 Word 更新时重跑本脚本生成新初版再人工校对。"""
import json, os, sys, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = r'c:\Users\lenovo\Desktop\题库'
DATA_DIR = os.path.join(BASE, 'data', 'tmp')
OUT = os.path.join(BASE, 'data', 'master_bank.xlsx')

HEADERS = ['章节', '源题号', '题型', '题干',
           '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H',
           '正确答案', '解析', '需复核']

JUDGE_MAP = {'对': '对', '是': '对', '正确': '对', '√': '对', '✓': '对', 'T': '对',
             '错': '错', '否': '错', '错误': '错', '×': '错', 'F': '错'}


def load_rows():
    rows = []
    for fn in ('parsed.json', 'bc_parsed.json'):
        data = json.load(open(os.path.join(DATA_DIR, fn), encoding='utf-8'))
        for g in data:
            ch = g['chapter']
            for q in g['questions']:
                if not q.get('valid'):
                    continue
                sec = q['sec_type']
                ans = (q['answer'] or '').strip()
                if sec == '判断题':
                    options = ['对', '错']
                    ans = JUDGE_MAP.get(ans, ans)
                else:
                    options = [t for _, t in q['options']][:8]
                # 选项列固定为 8 列，保证 正确答案/解析/需复核 与表头对齐
                options += [''] * (8 - len(options))
                rows.append([ch, q.get('src_no'), sec, q['stem'],
                             *options, ans, q.get('note') or '',
                             '；'.join(q.get('flags') or [])])
    return rows


def main():
    rows = load_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '母题库'
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)

    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(bold=True, color='FFFFFF')
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    for i in range(2, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=i, column=c).border = border
    for col, w in zip('ABCDEFGHIJKLMNO', [14, 7, 9, 46, 20, 20, 20, 20, 20, 20, 20, 20, 8, 28, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    wb.save(OUT)
    print('master_bank.xlsx 初版 ->', OUT, '共', len(rows), '题')


if __name__ == '__main__':
    main()
