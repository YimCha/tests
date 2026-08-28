# -*- coding: utf-8 -*-
"""一次性：把当前岗位配置（岗位分值、部门分组）写入 data/config.xlsx 初版。
之后岗位配置以 config.xlsx 为准，由 bank_loader.py 读取；本脚本仅用于初始化/重置。

config.xlsx 的"岗位配置"sheet 列说明（与《上岗考试实施方案》对齐）：
  岗位 | 类别(A/B/C) | 组 | 部门 | 分值
  - A 类行：组、部门留空，分值 = 该岗位 A 类总分
  - B/C 类行：每个部门一行，同组部门组号相同（表示共占比、组内混合抽题）；
    分值填在组内首行，同组其余行留空
"""
import os, sys, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = r'c:\Users\lenovo\Desktop\题库'
OUT = os.path.join(BASE, 'data', 'config.xlsx')

# 初始种子，与《上岗考试实施方案》一致（ratio 为题库内 80 分制分值）
POSITIONS = [
    {'name': '柜员岗',
     'ratio': [30, 40, 10],
     'b': [[['运营管理部', '数字金融部'], 40]],
     'c': [[['运营管理部'], 5], [['数字金融部'], 5]]},
    {'name': '客户经理岗',
     'ratio': [20, 50, 10],
     'b': [[['公司金融部', '个人金融部'], 10], [['风险管理部'], 10], [['授信审批部'], 10],
           [['普惠金融部', '资金营运中心'], 10], [['数字金融部'], 10]],
     'c': [[['公司金融部', '个人金融部', '风险管理部', '授信审批部', '普惠金融部', '资金营运中心', '数字金融部'], 10]]},
    {'name': '运营管理岗',
     'ratio': [20, 50, 10],
     'b': [[['运营管理部'], 25], [['计划财务部', '数据管理部'], 25]],
     'c': [[['运营管理部', '计划财务部', '数据管理部'], 10]]},
    {'name': '内控稽核岗',
     'ratio': [20, 50, 10],
     'b': [[['数据管理部'], 5], [['运营管理部'], 20], [['审计部'], 25]],
     'c': [[['运营管理部', '审计部'], 10]]},
    {'name': '科技岗',
     'ratio': [30, 40, 10],
     'b': [[['内控合规部'], 10], [['科技部'], 30]],
     'c': [[['安全保卫部', '内控合规部', '科技部'], 10]]},
    {'name': '行政管理岗',
     'ratio': [20, 50, 10],
     'b': [[['办公室'], 10], [['人力资源部'], 10], [['安全保卫部'], 10], [['数据管理部'], 10], [['内控合规部'], 10]],
     'c': [[['办公室', '人力资源部', '安全保卫部', '数据管理部', '内控合规部'], 10]]},
]

HEADER = ['岗位', '类别', '组', '部门', '分值']


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '岗位配置'
    ws.append(HEADER)
    for p in POSITIONS:
        ws.append([p['name'], 'A', '', '', p['ratio'][0]])
        for cls, groups in (('B', p['b']), ('C', p['c'])):
            for gid, (depts, pts) in enumerate(groups, 1):
                for i, d in enumerate(depts):
                    ws.append([p['name'], cls, gid, d, pts if i == 0 else ''])

    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(bold=True, color='FFFFFF')
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for i in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(row=i, column=c).border = border
    for col, w in zip('ABCDE', [12, 8, 6, 30, 8]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    wb.save(OUT)
    print('config.xlsx ->', OUT, '共', ws.max_row - 1, '行配置')


if __name__ == '__main__':
    main()
