# -*- coding: utf-8 -*-
"""从 master_bank.xlsx（经 bank_loader）按考试方案为 6 个岗位生成 ABC 类合集 Excel
（A类全部 + 该岗位B类 + 该岗位C类），并输出核对清单（每岗位题数/题型分布/逐题明细）。
岗位与部门比例来自 bank_loader 的 positions 配置；格式与 A类合并 保持一致（判断题统一 对/错 选项 + A/B 答案）。"""
import sys, io, os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bank_loader import load_bank

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
BASE = r'c:\Users\lenovo\Desktop\题库'
OUT_DIR = os.path.join(BASE, 'data', 'output', 'ABC类')

TYPE_NAME = {0: '单选题', 1: '多选题', 2: '判断题'}
JUDGE_MAP = {'对': 'A', '是': 'A', '正确': 'A', '√': 'A', '✓': 'A', 'T': 'A',
             '错': 'B', '否': 'B', '错误': 'B', '×': 'B', 'F': 'B'}

bank = load_bank()
chapters = bank['chapters']
POSITIONS = bank['positions']

TEMPLATE_HEADER = ['题干（必填）', '题型 （必填）',
                   '选项 A', '选项 B', '选项 C', '选项 D',
                   '选项E\n(勿删)', '选项F\n(勿删)', '选项G\n(勿删)', '选项H\n(勿删)',
                   '正确答案\n（必填）', '解析\n（勿删）', '章节\n（勿删）', '难度']


def collect_position(p):
    """返回 [(chapter, q), ...]，顺序：A类全部 -> B类 -> C类（部门按方案顺序）。"""
    rows = []
    for q in bank['q']:
        if chapters[q['ch']].endswith('A类'):
            rows.append((chapters[q['ch']], q))
    for cls, groups in (('B类', p['b']), ('C类', p['c'])):
        for depts, _ in groups:
            for dept in depts:
                ch = dept + cls
                for q in bank['q']:
                    if chapters[q['ch']] == ch:
                        rows.append((ch, q))
    return rows


def q_row(chapter, q):
    row = [None] * 14
    row[0] = q['s']
    row[1] = TYPE_NAME[q['t']]
    if q['t'] == 2:
        row[2], row[3] = '对', '错'
        row[10] = JUDGE_MAP.get(q['a'], q['a'])
    else:
        for i, text in enumerate(q['o']):
            if i < 8:
                row[2 + i] = text
        row[10] = q['a'] or ''
    row[11] = q['n'] or ''
    row[12] = chapter
    return row


def write_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = '试题案例，直接导入试试'
    ws.append(['导入须知\n1. 请一定要阅读本说明后，再编辑试题；导入时，本说明无需删除，可直接操作导入\n2. 第二行表头（蓝色那行）不能删除，否则无法导入'])
    ws.append(TEMPLATE_HEADER)
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(bold=True, color='FFFFFF')
    for c in range(1, 15):
        cell = ws.cell(row=2, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
    for chapter, q in rows:
        ws.append(q_row(chapter, q))
        r = ws.max_row
        for c in range(1, 15):
            ws.cell(row=r, column=c).border = border
    for col, w in zip('ABCDEFGHIJKLMN', [40, 10, 18, 18, 18, 18, 18, 18, 18, 18, 10, 30, 14, 8]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A3'
    wb.save(path)
    return ws.max_row - 2


def build_checklist(path, results):
    wb = Workbook()
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_font = Font(bold=True, color='FFFFFF')
    warn_fill = PatternFill('solid', fgColor='FFF2CC')
    warn_font = Font(color='9C6500')

    # Sheet1 汇总
    ws1 = wb.active
    ws1.title = '汇总'
    h1 = ['岗位', 'ABC总题数', 'A类题数', 'B类题数', 'C类题数', '单选题', '多选题', '判断题', '其他题型', '带标记需复核']
    ws1.append(h1)
    for name, rows in results:
        n = len(rows)
        b = sum(1 for ch, _ in rows if ch.endswith('B类'))
        c = sum(1 for ch, _ in rows if ch.endswith('C类'))
        st = Counter(TYPE_NAME[q['t']] for _, q in rows)
        other = sum(v for k, v in st.items() if k not in ('单选题', '多选题', '判断题'))
        flagged = sum(1 for _, q in rows if q['f'])
        ws1.append([name, n, n - b - c, b, c, st.get('单选题', 0), st.get('多选题', 0),
                    st.get('判断题', 0), other, flagged])
    for j in range(1, 11):
        ws1.cell(row=1, column=j).fill = hdr_fill
        ws1.cell(row=1, column=j).font = hdr_font
        ws1.cell(row=1, column=j).border = border
    for i in range(1, ws1.max_row + 1):
        for j in range(1, 11):
            ws1.cell(row=i, column=j).border = border
            if ws1.cell(row=i, column=10).value:
                ws1.cell(row=i, column=10).fill = warn_fill
                ws1.cell(row=i, column=10).font = warn_font
    for col, w in zip('ABCDEFGHIJ', [10, 11, 10, 10, 10, 9, 9, 9, 9, 13]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = 'A2'

    # Sheet2 逐题明细
    ws2 = wb.create_sheet('逐题明细')
    h2 = ['岗位', '序号', '章节', '题型', '答案', '题干', '解析/备注', '标记']
    ws2.append(h2)
    g = 0
    for name, rows in results:
        for chapter, q in rows:
            g += 1
            stem = ' '.join(q['s'].split())
            if len(stem) > 80:
                stem = stem[:80] + '…'
            ws2.append([name, g, chapter, TYPE_NAME[q['t']],
                        q['a'] if q['a'] else '',
                        stem, q['n'] if q['n'] else '',
                        '；'.join(q['f']) if q['f'] else ''])
    for j in range(1, 9):
        ws2.cell(row=1, column=j).fill = hdr_fill
        ws2.cell(row=1, column=j).font = hdr_font
        ws2.cell(row=1, column=j).border = border
    for i in range(2, ws2.max_row + 1):
        for j in range(1, 9):
            ws2.cell(row=i, column=j).border = border
        if ws2.cell(row=i, column=8).value:
            for j in range(1, 9):
                ws2.cell(row=i, column=j).fill = warn_fill
    for col, w in zip('ABCDEFGH', [10, 7, 14, 9, 9, 62, 30, 26]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    wb.save(path)
    return g


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    results = []
    for p in POSITIONS:
        rows = collect_position(p)
        path = os.path.join(OUT_DIR, f'{p["name"]}ABC类合集.xlsx')
        n = write_excel(path, rows)
        results.append((p['name'], rows))
        print(f'{p["name"]}: {n} 题 -> {os.path.basename(path)}')
    cl_path = os.path.join(OUT_DIR, '各岗位ABC类_核对清单.xlsx')
    tot = build_checklist(cl_path, results)
    print('核对清单 ->', os.path.basename(cl_path), '逐题总数', tot)


if __name__ == '__main__':
    main()
