# -*- coding: utf-8 -*-
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
wb = openpyxl.load_workbook(r'c:\Users\lenovo\Desktop\题库\kaoshibaoExcel20221101.xlsx')
print('sheets:', wb.sheetnames)
for ws in wb.worksheets:
    print('='*40)
    print('sheet:', ws.title, 'dims:', ws.dimensions, 'max_row:', ws.max_row, 'max_col:', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), max_col=min(ws.max_column, 12)):
        for c in row:
            if c.value is not None:
                print(c.coordinate, repr(c.value)[:80])
