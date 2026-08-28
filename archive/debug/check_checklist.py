# -*- coding: utf-8 -*-
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
wb = openpyxl.load_workbook(r'c:\Users\lenovo\Desktop\题库\A类题库_核对清单.xlsx')
print('sheets:', wb.sheetnames)
ws = wb['汇总']
print('--- 汇总 (前5行 + 合计) ---')
for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
    print(row)
last = list(ws.iter_rows(values_only=True))[-1]
print('合计行:', last)

ws2 = wb['逐题明细']
print('--- 逐题明细 行数:', ws2.max_row - 1)
for row in ws2.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row)
# 带标记的行
print('--- 带标记行 ---')
cnt = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[7]:
        cnt += 1
        print(row[0], row[1], row[2], row[3], row[4], '| 标记:', row[7])
print('标记总数:', cnt)
