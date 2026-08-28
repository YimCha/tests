# -*- coding: utf-8 -*-
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
wb = openpyxl.load_workbook(r'c:\Users\lenovo\Desktop\题库\kaoshibaoExcel20221101.xlsx')
ws = wb['试题案例，直接导入试试']
for col in range(13, 26):
    v = ws.cell(row=2, column=col).value
    v3 = ws.cell(row=3, column=col).value
    if v or v3:
        print(openpyxl.utils.get_column_letter(col), '| hdr:', repr(v), '| row3:', repr(v3))
# 合并单元格情况
print('merged:', ws.merged_cells.ranges)
