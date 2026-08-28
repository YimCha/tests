# -*- coding: utf-8 -*-
"""从 data/master_bank.xlsx（唯一人工编辑入口）读取题库、data/config.xlsx 读取岗位配置，
组装 BANK 结构。章节→岗位归属按章节名（XX部A/B/C类）自动推导；岗位分值、部门分组、
工具元数据（银行名称/标题/页脚/输出文件名）均由 config.xlsx 维护，不在代码中硬编码。"""
import os, re
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(BASE, 'data', 'master_bank.xlsx')
CONFIG = os.path.join(BASE, 'data', 'config.xlsx')

JUDGE_MAP = {'对': '对', '是': '对', '正确': '对', '√': '对', '✓': '对', 'T': '对',
             '错': '错', '否': '错', '错误': '错', '×': '错', 'F': '错'}
TYPE_ID = {'单选题': 0, '多选题': 1, '判断题': 2}
NUM_PREFIX = re.compile(r'^\d+\s*[.、．]')


def clean_stem(s):
    s = NUM_PREFIX.sub('', s or '').strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def load_rows():
    """返回 [(chapter, q), ...]，顺序按母题库行序。"""
    wb = openpyxl.load_workbook(MASTER)
    ws = wb['母题库'] if '母题库' in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {name: i for i, name in enumerate(hdr)}
    opts_cols = [idx[f'选项{c}'] for c in 'ABCDEFGH' if f'选项{c}' in idx]

    def col(r, name):
        i = idx.get(name)
        return r[i] if i is not None and i < len(r) else None

    out = []
    for r in it:
        chapter = col(r, '章节')
        sec = col(r, '题型')
        stem = col(r, '题干')
        if not (chapter and sec and stem):
            continue
        sec = str(sec).strip()
        tid = TYPE_ID.get(sec)
        if tid is None:
            continue
        ans = str(col(r, '正确答案') or '').strip()
        if tid == 2:
            ans = JUDGE_MAP.get(ans, ans)
        else:
            ans = re.sub(r'[^A-H]', '', ans.upper())
        options = [str(r[i] or '').strip() for i in opts_cols]
        options = ['对', '错'] if tid == 2 else [o for o in options if o]
        q = {
            't': tid,
            's': clean_stem(str(stem)),
            'o': options,
            'a': ans,
            'n': str(col(r, '解析') or '').strip(),
            'f': [x for x in str(col(r, '需复核') or '').split('；') if x],
        }
        out.append((str(chapter).strip(), q))
    return out


def load_positions():
    """从 config.xlsx 的'岗位配置'sheet 读取岗位配置，返回 [{name, ratio, b, c}]。
    列：岗位 | 类别(A/B/C) | 组 | 部门 | 分值。A 类行组/部门留空；B/C 类每个部门一行，
    同组部门组号相同（共占比），分值填组内首行。ratio=[A,B,C] 为题库内 80 分制分值；
    b/c 为部门组配置 [[[部门,...], 分值], ...]。"""
    wb = openpyxl.load_workbook(CONFIG)
    ws = wb['岗位配置'] if '岗位配置' in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    next(it)  # 表头
    positions = []
    for r in it:
        name = str(r[0]).strip() if r[0] else ''
        if not name:
            continue
        cls = str(r[1]).strip().upper() if r[1] else ''
        gid = int(r[2]) if r[2] is not None else None
        dept = str(r[3]).strip() if r[3] else ''
        pts = int(r[4]) if r[4] is not None else 0
        p = _get_pos(positions, name)
        if cls == 'A':
            p['ratio'][0] = pts
        elif cls in ('B', 'C'):
            if gid is None or not dept:
                raise ValueError(f'config.xlsx: {name} {cls} 类行缺少组号或部门: {r}')
            key = 1 if cls == 'B' else 2
            groups = p['b'] if key == 1 else p['c']
            while len(groups) < gid:
                groups.append([[], 0])
            g = groups[gid - 1]
            g[0].append(dept)
            if pts:
                g[1] = pts
            p['ratio'][key] += pts
        else:
            raise ValueError(f'config.xlsx: 未知类别 {r[1]!r}（岗位 {name}）')
    _check_positions(positions)
    return positions


def _get_pos(positions, name):
    for p in positions:
        if p['name'] == name:
            return p
    p = {'name': name, 'ratio': [0, 0, 0], 'b': [], 'c': []}
    positions.append(p)
    return p


def _check_positions(positions):
    if not positions:
        raise ValueError('config.xlsx 岗位配置为空')
    for p in positions:
        if p['ratio'][0] <= 0 or p['ratio'][1] <= 0 or p['ratio'][2] <= 0:
            raise ValueError(f"{p['name']}: A/B/C 分值配置不完整 {p['ratio']}")
        if sum(p['ratio']) != 80:
            raise ValueError(f"{p['name']}: A/B/C 分值合计 {sum(p['ratio'])} != 80")
        for cls, groups in (('B', p['b']), ('C', p['c'])):
            if not groups:
                raise ValueError(f"{p['name']}: 缺少 {cls} 类部门组配置")
            for depts, pts in groups:
                if not depts or pts <= 0:
                    raise ValueError(f"{p['name']} {cls} 类部门组配置无效: {depts} {pts}")


def _check_depts(positions, chapters):
    """校验配置中的部门在章节里真实存在，防止 Excel 手误产生静默错误归属。"""
    chapter_set = set(chapters)
    for p in positions:
        for cls, groups in (('B', p['b']), ('C', p['c'])):
            for depts, _ in groups:
                for d in depts:
                    ch = d + cls + '类'
                    if ch not in chapter_set:
                        raise ValueError(f"{p['name']} 配置部门 {d} 找不到章节 {ch}")


def positions_for(chapter, positions):
    if chapter.endswith('A类'):
        return list(range(len(positions)))
    cls = 'b' if chapter.endswith('B类') else ('c' if chapter.endswith('C类') else None)
    if cls is None:
        return []
    dept = chapter[:-2]
    return [i for i, p in enumerate(positions) if any(dept in g[0] for g in p[cls])]


def load_meta():
    """从 config.xlsx 的'工具配置'sheet 读取工具元数据（银行名称/标题/页脚/输出文件名）。
    列：键 | 值 | 说明。"""
    wb = openpyxl.load_workbook(CONFIG)
    if '工具配置' not in wb.sheetnames:
        return {}
    meta = {}
    for r in wb['工具配置'].iter_rows(values_only=True):
        if r and r[0] and str(r[0]).strip() != '键':
            meta[str(r[0]).strip()] = str(r[1] or '').strip()
    return meta


def load_bank():
    """返回 BANK 结构 {'v', 'positions', 'chapters', 'q', 'meta'}。"""
    positions = load_positions()
    rows = load_rows()
    chapters = []
    chap_idx = {}
    qs = []
    for chapter, q in rows:
        if chapter not in chap_idx:
            chap_idx[chapter] = len(chapters)
            chapters.append(chapter)
        q['ch'] = chap_idx[chapter]
        q['p'] = positions_for(chapter, positions)
        qs.append(q)
    _check_depts(positions, chapters)
    return {
        'v': 2,
        'positions': positions,
        'chapters': chapters,
        'q': qs,
        'meta': load_meta(),
    }


if __name__ == '__main__':
    import io, sys
    from collections import Counter
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    bank = load_bank()
    print('master_bank.xlsx ->', MASTER, '| config.xlsx ->', CONFIG)
    print('岗位数:', len(bank['positions']), '| 章节数:', len(bank['chapters']), '| 题目数:', len(bank['q']))
    print('题型分布:', dict(Counter(q['t'] for q in bank['q'])))
    for i, p in enumerate(bank['positions']):
        n = sum(1 for q in bank['q'] if i in q['p'])
        b = sum(1 for q in bank['q'] if i in q['p'] and bank['chapters'][q['ch']].endswith('B类'))
        c = sum(1 for q in bank['q'] if i in q['p'] and bank['chapters'][q['ch']].endswith('C类'))
        print(f'  {p["name"]}: 合计{n} (B类{b} C类{c})')
